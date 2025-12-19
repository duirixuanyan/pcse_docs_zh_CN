# -*- coding: utf-8 -*-
# 版权所有 (c) 2004-2024 Wageningen Environmental Research, Wageningen-UR
# Allard de Wit (allard.dewit@wur.nl), 2024年3月
"""从Excel文件读取天气数据的数据提供者。
"""
import os
import openpyxl

from ..base import WeatherDataContainer, WeatherDataProvider
from ..util import reference_ET, angstrom, check_angstromAB
from ..exceptions import PCSEError
from ..settings import settings

# 转换函数
NoConversion = lambda x: x
kJ_to_J = lambda x: x*1000.
kPa_to_hPa = lambda x: x*10.
mm_to_cm = lambda x: x/10.


def determine_true_false(value):
    """OpenPyXL对true/false有点奇怪的处理方式

    Excel单元格值         OpenPyXL单元格值            类型
       =FALSE()            '=FALSE()'               str
       =FALSE              '=FALSE'                 str
       false                False                   bool
       =TRUE()             '=TRUE()'                str
       =TRUE               '=TRUE'                  str
       TRUE                 True                    bool

    最后，true/false也有可能以0/1整数值表示。

    该函数尝试处理所有这些情况。
    """
    if isinstance(value, str):
        v = value.lower()
        if "true" in v:
            return True
        elif "false" in v:
            return False
    elif isinstance(value, bool):
        return value
    elif isinstance(value, int):
        if value == 1:
            return True
        elif value == 0:
            return False

    msg = f"cannot determine True|False: {value}"
    raise ValueError(msg)


class ExcelWeatherDataProvider(WeatherDataProvider):
    """从 Excel 文件（仅 .xlsx 格式）读取天气数据。

    :param xls_fname: 要读取的 Excel 文件名称
    :param mising_snow_depth: 缺失 SNOW_DEPTH 值时应使用的值，默认为 None。
    :param force_reload: 是否跳过缓存文件，从 .xlsx 文件重新加载数据并写入新的缓存文件。缓存文件保存于 `$HOME/.pcse/meteo_cache`

    早期仅有 CABOWeatherDataProvider 可用于从文件读取天气数据，其数据来源于 CABO 天气格式的文本文件。
    然而，生成 CABO 天气文件非常繁琐，因为每年必须构建一个新文件，并且容易出错，格式稍有问题就导致错误。

    为简化向 PCSE 模型提供天气数据的流程，开发了一个新的数据提供类，该类可直接从简单的 excel 文件中读取数据。

    ExcelWeatherDataProvider 假定数据记录是完整的，不会尝试对数据进行插值，因为这在 Excel 中可轻松完成。
    只有 SNOW_DEPTH 可以缺失，因为该参数通常在非冬季不提供。
    """
    obs_conversions = {
        "TMAX": NoConversion,
        "TMIN": NoConversion,
        "IRRAD": kJ_to_J,
        "VAP": kPa_to_hPa,
        "WIND": NoConversion,
        "RAIN": mm_to_cm,
        "SNOWDEPTH": NoConversion
    }
    # 允许读取的最大行数，假设最多有 250 年的天气数据。
    # 这样做的原因是，在某些 Excel 文件中，sheet.max_row 的属性值为 Excel 工作表允许的最大行数，
    # 这会导致读取到数百万行的空记录。
    max_rows = int(250*365.35)

    # 各类值开始的行号。注意行号为零起，因此要加 1 才是 Excel 中的实际行号。
    site_row = 9
    label_row = 11
    data_start_row = 13

    def __init__(self, xls_fname, missing_snow_depth=None, force_reload=False):
        # 初始化，读取天气数据提供文件路径和相关参数
        WeatherDataProvider.__init__(self)

        self.fp_xls_fname = os.path.abspath(xls_fname)
        self.missing_snow_depth = missing_snow_depth
        if not os.path.exists(self.fp_xls_fname):
            msg = "Cannot find weather file at: %s" % self.fp_xls_fname
            raise PCSEError(msg)

        # 如果强制重载，或未能成功加载缓存，则从excel文件读取数据
        if force_reload or not self._load_cache_file(self.fp_xls_fname):
            book = openpyxl.load_workbook(self.fp_xls_fname, read_only=True)
            sheet = book.active

            self._read_header(sheet)  # 读取头部信息
            self._read_site_characteristics(sheet)  # 读取站点特征信息
            self._read_observations(sheet)  # 读取观测数据

            self._write_cache_file(self.fp_xls_fname)  # 写入缓存文件

    def _read_header(self, sheet):
        # 读取数据头部信息，包括国家、站点、描述、来源、联系人等
        country = sheet["B2"].value
        station = sheet["B3"].value
        desc = sheet["B4"].value
        src = sheet["B5"].value
        contact = sheet["B6"].value
        self.nodata_value = float(sheet["B7"].value)
        self.description = [u"Weather data for:",
                            u"Country: %s" % country,
                            u"Station: %s" % station,
                            u"Description: %s" % desc,
                            u"Source: %s" % src,
                            u"Contact: %s" % contact]

    def _read_site_characteristics(self, sheet):
        # 读取站点的地理信息和安格斯特朗系数等
        self.longitude = float(sheet[f"A{self.site_row}"].value)
        self.latitude = float(sheet[f"B{self.site_row}"].value)
        self.elevation = float(sheet[f"C{self.site_row}"].value)
        angstA = float(sheet[f"D{self.site_row}"].value)
        angstB = float(sheet[f"E{self.site_row}"].value)
        self.angstA, self.angstB = check_angstromAB(angstA, angstB)
        try:
            has_sunshine = sheet[f"F{self.site_row}"].value
            self.has_sunshine = determine_true_false(has_sunshine)
        except ValueError as e:
            # 如果无法判断是否包含辐射或日照时数，抛出异常
            raise PCSEError(f"Cannot determine if sheet as radiation or sunshine hours: {e}")


    def _read_observations(self, sheet):
        # 首先获取列标签
        labels = [cell.value for cell in sheet[self.label_row]]

        # 开始读取有数据的所有行
        max_row = min(sheet.max_row, self.max_rows)
        for rownum, row in enumerate(sheet[self.data_start_row:max_row]):
            try:
                d = {}
                for cell, label in zip(row, labels):
                    if label == "DAY":
                        if cell.value is None:
                            raise ValueError
                        else:
                            d[label] = cell.value.date()
                            continue

                    # 显式转换为浮点数。如果失败，将抛出ValueError
                    value = float(cell.value)

                    # 检查观测值是否被标记为缺失。目前只有SNOWDEPTH允许有缺失值。否则抛出错误
                    if self._is_missing_value(value):
                        if label == "SNOWDEPTH":
                            value = self.missing_snow_depth
                        else:
                            raise ValueError()

                    if label == "IRRAD" and self.has_sunshine is True:
                        if 0 <= value <= 24:
                            # 使用Angstrom方程将日照时数转换为辐射（J/m2/day）
                            value = angstrom(d["DAY"], self.latitude, value, self.angstA, self.angstB)
                            value /= 1000.  # 转换为kJ/m2/day，以便与obs_conversion函数兼容
                        else:
                            msg = "Sunshine duration not within 0-24 interval for row %i" % \
                                  (rownum + self.data_start_row)
                            raise ValueError(msg)

                    func = self.obs_conversions[label]
                    d[label] = func(value)

                # 参考作物蒸散发（mm/day）
                e0, es0, et0 = reference_ET(LAT=self.latitude, ELEV=self.elevation, ANGSTA=self.angstA,
                                            ANGSTB=self.angstB, **d)
                # 转换为cm/day
                d["E0"] = e0/10.; d["ES0"] = es0/10.; d["ET0"] = et0/10.

                wdc = WeatherDataContainer(LAT=self.latitude, LON=self.longitude, ELEV=self.elevation, **d)
                self._store_WeatherDataContainer(wdc, d["DAY"])

            except ValueError as e:  # 单元格中有异常值
                msg = "Failed reading row: %i. Skipping..." % (rownum + self.data_start_row)
                self.logger.warning(msg)
                print(msg)

    def _load_cache_file(self, xls_fname):
         # 尝试查找缓存文件，如果不存在则返回False，否则加载
         cache_filename = self._find_cache_file(xls_fname)
         if cache_filename is None:
             return False
         else:
             try:
                 self._load(cache_filename)
                 return True
             except:
                 return False

    def _find_cache_file(self, xls_fname):
        """尝试查找xls_fname对应的缓存文件

        如果缓存文件不存在，返回None；否则返回缓存文件的完整路径。
        """
        cache_filename = self._get_cache_filename(xls_fname)
        if os.path.exists(cache_filename):
            cache_date = os.stat(cache_filename).st_mtime
            xls_date = os.stat(xls_fname).st_mtime
            if cache_date > xls_date:  # 缓存要比XLS文件更新
                return cache_filename

        return None

    def _get_cache_filename(self, xls_fname):
        """根据给定的xls_fname构建用于缓存文件的文件名
        """
        basename = os.path.basename(xls_fname)
        filename, ext = os.path.splitext(basename)

        tmp = "%s_%s.cache" % (self.__class__.__name__, filename)
        cache_filename = os.path.join(settings.METEO_CACHE_DIR, tmp)
        return cache_filename

    def _write_cache_file(self, xls_fname):
        # 写缓存文件
        cache_filename = self._get_cache_filename(xls_fname)
        try:
            self._dump(cache_filename)
        except (IOError, EnvironmentError) as e:
            msg = "Failed to write cache to file '%s' due to: %s" % (cache_filename, e)
            self.logger.warning(msg)

    def _is_missing_value(self, value):
        """检查该值是否等于指定的缺失值

        :return: True|False
        """
        eps = 0.0001
        if abs(value - self.nodata_value) < eps:
            return True
        else:
            return False